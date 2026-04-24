"""Run Victorinox file with improvements and compare found rate."""
import io
import os
import sys
import time

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import httpx
import openpyxl

base = "http://127.0.0.1:8000"

# Reconstruct original from the results file (strip Trasco columns)
results_path = r"c:\Users\Avi\Downloads\2026-03-26 VICTORINOX SWISS MILITARY (1)_trasco_results.xlsx"
wb = openpyxl.load_workbook(results_path)
ws = wb[wb.sheetnames[0]]
headers = [str(ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1)]
trasco_cols = {"ASIN", "confidence", "Trasco trace", "Rejected ASIN (LLM)"}
dbg_prefix = "dbg_"
orig_cols = [
    (i + 1, h)
    for i, h in enumerate(headers)
    if h not in trasco_cols and not h.startswith(dbg_prefix)
]
new_wb = openpyxl.Workbook()
new_ws = new_wb.active
new_ws.title = wb.sheetnames[0]
for ci, (src_col, h) in enumerate(orig_cols, start=1):
    new_ws.cell(1, ci, h)
for r in range(2, ws.max_row + 1):
    for ci, (src_col, _) in enumerate(orig_cols, start=1):
        new_ws.cell(r, ci, ws.cell(r, src_col).value)
buf = io.BytesIO()
new_wb.save(buf)
data = buf.getvalue()
orig_name = "2026-03-26 VICTORINOX SWISS MILITARY.xlsx"
print(f"Reconstructed original ({len(orig_cols)} columns, {ws.max_row - 1} rows)")

# Submit job
r = httpx.post(
    f"{base}/api/v1/process/start?use_ollama=false&use_ollama_asin_validate=true&debug=true",
    files={"file": (orig_name, data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    timeout=30,
)
print(f"Start: {r.status_code} {r.json()}")
job_id = r.json().get("job_id")
if not job_id:
    sys.exit(1)

# Poll
while True:
    time.sleep(3)
    sr = httpx.get(f"{base}/api/v1/process/status/{job_id}", timeout=30)
    s = sr.json()
    phase = s.get("phase", "")
    msg = s.get("message", "")[:80]
    status = s.get("status", "")
    elapsed = s.get("elapsed_sec", 0)
    print(f"  [{elapsed:6.1f}s] {phase:20s} {msg}")
    if status in ("complete", "error"):
        print(f"Final: {status}")
        if status == "error":
            print(f"Error: {s.get('error')}")
            sys.exit(1)
        break

# Download
dr = httpx.get(f"{base}/api/v1/process/result/{job_id}", timeout=60)
outpath = r"c:\Users\Avi\Downloads\victorinox_improved_test.xlsx"
with open(outpath, "wb") as f:
    f.write(dr.content)
print(f"Saved to {outpath}")

# Analyze
wb2 = openpyxl.load_workbook(outpath, data_only=True)
ws2 = wb2[wb2.sheetnames[0]]
h2 = [str(ws2.cell(1, c).value or "") for c in range(1, ws2.max_column + 1)]
idx = {name: i + 1 for i, name in enumerate(h2)}

total = found = 0
confs = {}
for r in range(2, ws2.max_row + 1):
    conf = str(ws2.cell(r, idx["confidence"]).value or "").strip()
    if not conf:
        continue
    total += 1
    confs[conf] = confs.get(conf, 0) + 1
    asin = str(ws2.cell(r, idx["ASIN"]).value or "").strip()
    if asin:
        found += 1

print(f"\n{'='*50}")
print(f"RESULTS: {found}/{total} ASINs found ({100*found/max(total,1):.1f}%)")
print(f"(Previous: 381/530 = 71.9%)")
print(f"\nConfidence breakdown:")
for k, v in sorted(confs.items(), key=lambda x: -x[1]):
    print(f"  {k:25s} {v:4d} ({100*v/total:.1f}%)")

# Brand inference check
brand_col = idx.get("dbg_parsed_brand")
if brand_col:
    brands = set()
    for r in range(2, min(10, ws2.max_row + 1)):
        b = str(ws2.cell(r, brand_col).value or "").strip()
        if b:
            brands.add(b)
    print(f"\nInferred brands: {brands}")

# Fallback stats
if "dbg_fallback_attempted" in idx:
    fb_stats = {}
    for r in range(2, ws2.max_row + 1):
        fb = str(ws2.cell(r, idx["dbg_fallback_attempted"]).value or "").strip()
        if fb:
            fbr = str(ws2.cell(r, idx["dbg_fallback_result"]).value or "").strip()[:30]
            key = f"{fb}|{fbr}"
            fb_stats[key] = fb_stats.get(key, 0) + 1
    if fb_stats:
        print(f"\nFallback stats:")
        for k, v in sorted(fb_stats.items(), key=lambda x: -x[1])[:15]:
            print(f"  {k:50s} {v:4d}")
