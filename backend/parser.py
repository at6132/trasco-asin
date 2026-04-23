"""
Spreadsheet ingestion: multi-sheet xlsx, CSV, header detection, semantic columns.
Uses Claude Haiku when configured; optional local Ollama as fallback — not ASIN-centric.
"""

from __future__ import annotations

import csv
import io
import json
import logging
import re
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any, Optional

import httpx
from openpyxl import load_workbook

from backend.anthropic_usage import AnthropicUsageLedger, record_anthropic_messages_response
from backend.ollama_usage import OllamaTokenLedger, record_chat_response

if __package__:
    from .validator import normalize_asin, normalize_gtin
else:
    from validator import normalize_asin, normalize_gtin

logger = logging.getLogger(__name__)

SEMANTIC_FIELDS = ("asin", "ean", "upc", "mpn", "sku", "title", "description", "brand")

HEADER_KEYWORDS = frozenset(
    (
        "sku",
        "asin",
        "amazon",
        "ean",
        "upc",
        "gtin",
        "jan",
        "barcode",
        "isbn",
        "mpn",
        "part",
        "model",
        "article",
        "item",
        "product",
        "description",
        "title",
        "brand",
        "manufacturer",
        "price",
        "cost",
        "msrp",
        "net",
        "qty",
        "quantity",
        "pack",
        "packaging",
        "coo",
        "warehouse",
        "dropship",
        "hierarchy",
        "segment",
        "series",
        "application",
        "material",
        "color",
        "size",
        "shape",
        "sap",
        "ecc",
        "code",
        "order",
        "subtotal",
        "delivery",
        "customer",
    )
)


@dataclass
class ParseResult:
    headers: list[str]
    header_row_1based: int
    mapping: dict[str, Optional[str]]
    rows: list[dict[str, Any]]
    detection_method: str
    sheets_processed: list[str] = field(default_factory=list)


def _cell_str(v: Any, max_len: int = 120) -> str:
    if v is None:
        return ""
    s = str(v).strip().replace("\n", " ")
    return s[:max_len] if s else ""


def _headers_from_row(row: list[Any], max_cols: int = 64) -> list[str]:
    headers: list[str] = []
    for i in range(min(len(row), max_cols)):
        h = row[i]
        if h is None or str(h).strip() == "":
            headers.append(f"column_{i + 1}")
        else:
            headers.append(str(h).strip())
    return headers


def _score_header_candidate(row: list[Any]) -> float:
    cells: list[Any] = []
    for i in range(min(len(row), 48)):
        v = row[i]
        if v is None:
            continue
        if isinstance(v, str) and not v.strip():
            continue
        cells.append(v)
    if len(cells) < 2:
        return 0.0

    joined = " ".join(_cell_str(c, 400).lower() for c in cells)
    hits = sum(1 for kw in HEADER_KEYWORDS if kw in joined)

    if len(cells) == 1 and len(joined) > 140:
        return 0.0

    textish = 0.0
    for c in cells:
        t = _cell_str(c, 200)
        if not t:
            continue
        if len(t) > 100:
            continue
        if re.fullmatch(r"-?\d+(\.\d+)?", t):
            continue
        if re.fullmatch(r"\d{8,14}", t):
            continue
        textish += 1.0

    ratio = textish / max(len(cells), 1)
    if hits == 0 and ratio < 0.45:
        return 0.0
    return hits * 3.0 + ratio * 2.0 + min(len(cells), 20) * 0.08


def guess_header_row_index(rows: list[list[Any]], max_scan: int = 120) -> int:
    best_i, best_s = 0, -1.0
    lim = min(len(rows), max_scan)
    for i in range(lim):
        s = _score_header_candidate(rows[i])
        if s > best_s:
            best_i, best_s = i, s
    if best_s <= 0:
        return 0
    return best_i


def _header_strength(headers: list[str]) -> int:
    return sum(1 for h in headers if h and not h.startswith("column_"))


def _heuristic_mapping(headers: list[str]) -> dict[str, Optional[str]]:
    lower = {h.lower().strip(): h for h in headers}

    def _word_boundary_match(pattern: str, text: str) -> bool:
        return bool(re.search(r"(?:^|[\s_/\-\.])(" + re.escape(pattern) + r")(?:[\s_/\-\.]|$)", text))

    def pick_exact_then_boundary(*candidates: str) -> Optional[str]:
        for c in candidates:
            if c in lower:
                return lower[c]
        for c in candidates:
            for key, orig in lower.items():
                if _word_boundary_match(c, key):
                    return orig
        return None

    return {
        "asin": pick_exact_then_boundary("asin", "amazon asin", "amazon id"),
        "ean": pick_exact_then_boundary("ean", "gtin", "barcode", "jan code", "ean/upc", "upc/ean"),
        "upc": pick_exact_then_boundary("upc", "upc code"),
        "mpn": pick_exact_then_boundary(
            "mpn", "mfr part", "manufacturer item", "manufacturer part",
            "model number", "model no", "part number", "part no",
            "article number", "article no", "item code", "item number",
        ),
        "sku": pick_exact_then_boundary(
            "sku", "merchant sku", "seller sku", "vendor sku",
            "product code", "article code", "item sku",
        ),
        "title": pick_exact_then_boundary(
            "title", "product name", "product title",
            "short description", "long description", "item name",
        ),
        "description": pick_exact_then_boundary("description", "product description", "desc"),
        "brand": pick_exact_then_boundary("brand", "manufacturer", "mfr name", "brand name"),
    }


def _merge_mapping_prefer_primary(
    primary: dict[str, Optional[str]],
    fallback: dict[str, Optional[str]],
) -> dict[str, Optional[str]]:
    out: dict[str, Optional[str]] = {}
    for k in SEMANTIC_FIELDS:
        pv, fv = primary.get(k), fallback.get(k)
        out[k] = pv if pv else fv
    return out


def _mapping_has_any_column(m: dict[str, Optional[str]]) -> bool:
    return any(m.get(k) for k in SEMANTIC_FIELDS)


def _parse_json_object_from_model(content: str) -> dict[str, Any]:
    text = (content or "").strip()
    if not text:
        return {}
    fence = re.search(r"```(?:json)?\s*([\s\S]*?)\s*```", text, re.IGNORECASE)
    if fence:
        text = fence.group(1).strip()
    try:
        return json.loads(text) if text else {}
    except json.JSONDecodeError:
        pass
    start = text.find("{")
    end = text.rfind("}")
    if start >= 0 and end > start:
        try:
            return json.loads(text[start : end + 1])
        except json.JSONDecodeError:
            return {}
    return {}


def _matrix_preview_for_llm(rows: list[list[Any]], max_rows: int = 40, max_cols: int = 22) -> list[list[str]]:
    out: list[list[str]] = []
    for r in rows[:max_rows]:
        row_out: list[str] = []
        for c in r[:max_cols]:
            row_out.append(_cell_str(c, 100))
        out.append(row_out)
    return out


def _llm_prompt_block(
    rows_preview: list[list[Any]],
    guessed_header_1based: int,
    headers_if_guess: list[str],
) -> tuple[str, dict[str, Any]]:
    grid = _matrix_preview_for_llm(rows_preview)
    schema = {
        "header_row_1based": "integer|null",
        "columns": {k: "string|null" for k in SEMANTIC_FIELDS},
    }
    prompt = (
        "Spreadsheets are vendor price lists or offers. Row 1 is NOT always the table header — "
        "there may be logos, titles, blank rows, or key/value metadata before the real column header row.\n"
        "Given a preview grid (0-based rows), choose the ONE row that is the product table's COLUMN HEADERS "
        "(short labels like SKU, EAN, Description, Brand, Price — not a data row, not a paragraph).\n"
        "Then map headers to semantic keys. Values in 'columns' MUST be the EXACT header string from that row, "
        "or null if the sheet truly has no such column. Amazon ASIN is rare; leave asin null unless an ASIN column exists.\n"
        f"Heuristic guess for header row (1-based index): {guessed_header_1based}\n"
        f"If that row looks correct, you may repeat it or correct it.\n"
        f"Headers from heuristic guess: {json.dumps(headers_if_guess)}\n"
        f"Grid (row-major, truncated cells): {json.dumps(grid)}\n"
        f"Return ONLY JSON matching this shape: {json.dumps(schema)}"
    )
    return prompt, schema


def _parse_llm_sheet_response(parsed: dict[str, Any]) -> tuple[Optional[int], dict[str, Optional[str]]]:
    hr: Optional[int] = None
    raw_h = parsed.get("header_row_1based")
    if raw_h is None:
        raw_h = parsed.get("header_row")
    if isinstance(raw_h, (int, float)):
        hr = int(raw_h)
    elif isinstance(raw_h, str) and raw_h.strip().isdigit():
        hr = int(raw_h.strip())

    cols_in = parsed.get("columns")
    if not isinstance(cols_in, dict):
        cols_in = {k: parsed.get(k) for k in SEMANTIC_FIELDS}

    out_map: dict[str, Optional[str]] = {}
    for k in SEMANTIC_FIELDS:
        v = cols_in.get(k)
        if v is None or v == "":
            out_map[k] = None
            continue
        vs = str(v).strip()
        out_map[k] = vs if vs else None
    return hr, out_map


def _ollama_sheet_understanding(
    rows_preview: list[list[Any]],
    guessed_header_1based: int,
    headers_if_guess: list[str],
    *,
    ollama_base_url: str,
    ollama_model: str,
    timeout: float = 240.0,
    ollama_usage: Optional[OllamaTokenLedger] = None,
) -> tuple[Optional[int], dict[str, Optional[str]]]:
    prompt, _schema = _llm_prompt_block(rows_preview, guessed_header_1based, headers_if_guess)
    url = ollama_base_url.rstrip("/") + "/api/chat"
    payload = {
        "model": ollama_model,
        "messages": [{"role": "user", "content": prompt}],
        "stream": False,
        "format": "json",
    }
    with httpx.Client(timeout=timeout) as client:
        r = client.post(url, json=payload)
        r.raise_for_status()
        data = r.json()
        record_chat_response(ollama_usage, data)
    content = (data.get("message") or {}).get("content") or ""
    parsed = _parse_json_object_from_model(content)
    return _parse_llm_sheet_response(parsed)


def _ollama_reachable(base_url: str, timeout: float = 3.0) -> bool:
    try:
        with httpx.Client(timeout=timeout) as client:
            r = client.get(base_url.rstrip("/") + "/api/tags")
            return r.status_code == 200
    except Exception:
        return False


def _anthropic_sheet_understanding(
    rows_preview: list[list[Any]],
    guessed_header_1based: int,
    headers_if_guess: list[str],
    *,
    api_key: str,
    model: str,
    timeout: float = 120.0,
    anthropic_usage: Optional[AnthropicUsageLedger] = None,
) -> tuple[Optional[int], dict[str, Optional[str]]]:
    prompt, _schema = _llm_prompt_block(rows_preview, guessed_header_1based, headers_if_guess)
    url = "https://api.anthropic.com/v1/messages"
    body = {
        "model": model,
        "max_tokens": 2048,
        "messages": [{"role": "user", "content": prompt}],
    }
    headers = {
        "x-api-key": api_key,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json",
    }
    with httpx.Client(timeout=timeout) as client:
        r = client.post(url, json=body, headers=headers)
        if r.status_code >= 400:
            snippet = (r.text or "").strip().replace("\n", " ")[:600]
            raise RuntimeError(f"Anthropic HTTP {r.status_code}: {snippet or r.reason_phrase}")
        data = r.json()
        record_anthropic_messages_response(anthropic_usage, data)
    parts = data.get("content") or []
    text = ""
    for p in parts:
        if isinstance(p, dict) and p.get("type") == "text":
            text += p.get("text") or ""
    parsed = _parse_json_object_from_model(text)
    return _parse_llm_sheet_response(parsed)


def _validate_mapping_against_headers(
    mapping: dict[str, Optional[str]], headers: list[str]
) -> dict[str, Optional[str]]:
    hset = set(headers)
    return {k: (v if v in hset else None) for k, v in mapping.items()}


def _sheet_title_for_match(row: dict[str, Any], mapping: dict[str, Optional[str]]) -> Optional[str]:
    tcol, dcol = mapping.get("title"), mapping.get("description")
    if tcol:
        v = row.get(tcol)
        if v is not None and str(v).strip():
            return str(v).strip()
    if dcol:
        v = row.get(dcol)
        if v is not None and str(v).strip():
            return str(v).strip()
    return None


def _sheet_brand_text(row: dict[str, Any], mapping: dict[str, Optional[str]]) -> Optional[str]:
    bcol = mapping.get("brand")
    if not bcol:
        return None
    v = row.get(bcol)
    if v is None or not str(v).strip():
        return None
    return str(v).strip()


def _parse_sheet_matrix(
    raw: list[list[Any]],
    *,
    sheet_name: str,
    use_llm: bool,
    allow_ollama: bool,
    ollama_base_url: str,
    ollama_model: str,
    anthropic_api_key: str,
    haiku_model: str,
    timeout: float,
    ollama_usage: Optional[OllamaTokenLedger] = None,
    anthropic_usage: Optional[AnthropicUsageLedger] = None,
) -> ParseResult:
    if not raw:
        return ParseResult(
            headers=[],
            header_row_1based=1,
            mapping={k: None for k in SEMANTIC_FIELDS},
            rows=[],
            detection_method="heuristic",
            sheets_processed=[sheet_name],
        )

    guess_idx = guess_header_row_index(raw)
    guess_headers = _headers_from_row(raw[guess_idx])

    ollama_header_1based: Optional[int] = None
    ollama_cols: dict[str, Optional[str]] = {k: None for k in SEMANTIC_FIELDS}
    method = "heuristic"

    if use_llm:
        preview = raw[: min(len(raw), 80)]
        llm_ok = False
        if anthropic_api_key.strip():
            try:
                ollama_header_1based, ollama_cols = _anthropic_sheet_understanding(
                    preview,
                    guessed_header_1based=guess_idx + 1,
                    headers_if_guess=guess_headers,
                    api_key=anthropic_api_key.strip(),
                    model=haiku_model or "claude-haiku-4-5-20251001",
                    timeout=min(timeout, 120.0),
                    anthropic_usage=anthropic_usage,
                )
                method = "haiku"
                llm_ok = True
            except Exception as e:
                logger.warning("Anthropic Haiku sheet understanding failed: %s", e)
        if not llm_ok and allow_ollama and _ollama_reachable(ollama_base_url):
            try:
                ollama_header_1based, ollama_cols = _ollama_sheet_understanding(
                    preview,
                    guessed_header_1based=guess_idx + 1,
                    headers_if_guess=guess_headers,
                    ollama_base_url=ollama_base_url,
                    ollama_model=ollama_model,
                    timeout=timeout,
                    ollama_usage=ollama_usage,
                )
                llm_ok = True
                method = "ollama"
            except Exception as e:
                logger.warning("Ollama sheet understanding failed: %s", e)
        if not llm_ok:
            ollama_cols = {k: None for k in SEMANTIC_FIELDS}

    header_idx = guess_idx
    if ollama_header_1based is not None:
        h0 = ollama_header_1based - 1
        if 0 <= h0 < len(raw):
            cand = _headers_from_row(raw[h0])
            if _header_strength(cand) >= 2:
                header_idx = h0

    headers = _headers_from_row(raw[header_idx])
    body = raw[header_idx + 1 :]
    body = [list(r) for r in body if r and any(v is not None and str(v).strip() != "" for v in r)]

    heuristic = _heuristic_mapping(headers)
    if method in ("ollama", "haiku"):
        ollama_cols = _validate_mapping_against_headers(ollama_cols, headers)
        if _mapping_has_any_column(ollama_cols):
            mapping = _merge_mapping_prefer_primary(ollama_cols, heuristic)
        else:
            mapping = dict(heuristic)
            method = "heuristic"
    else:
        mapping = dict(heuristic)

    asin_col = mapping.get("asin")
    ean_col = mapping.get("ean")
    upc_col = mapping.get("upc")
    sku_col = mapping.get("sku")
    mpn_col = mapping.get("mpn")
    brand_col = mapping.get("brand")

    rows: list[dict[str, Any]] = []
    for bi, raw_row in enumerate(body):
        sheet_row = header_idx + 2 + bi
        row_obj: dict[str, Any] = {"_row_index": sheet_row, "_sheet_name": sheet_name}
        for i, h in enumerate(headers):
            row_obj[h] = raw_row[i] if i < len(raw_row) else None

        row_obj["_asin"] = normalize_asin(row_obj.get(asin_col)) if asin_col else None
        gtin: Optional[str] = None
        if ean_col:
            gtin = normalize_gtin(row_obj.get(ean_col)) or gtin
        if not gtin and upc_col:
            gtin = normalize_gtin(row_obj.get(upc_col))
        row_obj["_gtin"] = gtin

        if sku_col:
            v = row_obj.get(sku_col)
            row_obj["_sku"] = str(v).strip() if v is not None and str(v).strip() else None
        else:
            row_obj["_sku"] = None
        if mpn_col:
            v = row_obj.get(mpn_col)
            row_obj["_mpn"] = str(v).strip() if v is not None and str(v).strip() else None
        else:
            row_obj["_mpn"] = None

        row_obj["_sheet_title_text"] = _sheet_title_for_match(row_obj, mapping)
        row_obj["_sheet_brand"] = _sheet_brand_text(row_obj, mapping)
        row_obj["_mapping"] = dict(mapping)
        row_obj["_column_order"] = list(headers)
        rows.append(row_obj)

    return ParseResult(
        headers=headers,
        header_row_1based=header_idx + 1,
        mapping=mapping,
        rows=rows,
        detection_method=method,
        sheets_processed=[sheet_name],
    )


def _estimate_data_rows(raw: list[list[Any]]) -> int:
    if not raw:
        return 0
    idx = guess_header_row_index(raw)
    return max(0, len(raw) - idx - 1)


def parse_uploaded_file(
    data: bytes,
    filename: str,
    *,
    ollama_base_url: str = "http://127.0.0.1:11434",
    ollama_model: str = "gemma3:27b",
    use_ollama: bool = True,
    timeout: float = 240.0,
    anthropic_api_key: str = "",
    haiku_model: str = "claude-haiku-4-5-20251001",
    ollama_usage: Optional[OllamaTokenLedger] = None,
    anthropic_usage: Optional[AnthropicUsageLedger] = None,
) -> ParseResult:
    fn = (filename or "upload").lower()
    if fn.endswith(".csv"):
        use_llm_sheet = bool(use_ollama) or bool(anthropic_api_key.strip())
        return _parse_csv_bytes(
            data,
            sheet_name="csv",
            use_llm=use_llm_sheet,
            allow_ollama=bool(use_ollama),
            ollama_base_url=ollama_base_url,
            ollama_model=ollama_model,
            anthropic_api_key=anthropic_api_key,
            haiku_model=haiku_model,
            timeout=timeout,
            ollama_usage=ollama_usage,
            anthropic_usage=anthropic_usage,
        )

    wb = load_workbook(filename=BytesIO(data), read_only=True, data_only=True)
    try:
        names = list(wb.sheetnames)
        if not names:
            return ParseResult(
                headers=[],
                header_row_1based=1,
                mapping={k: None for k in SEMANTIC_FIELDS},
                rows=[],
                detection_method="heuristic",
                sheets_processed=[],
            )

        estimates: list[tuple[str, int]] = []
        matrices: dict[str, list[list[Any]]] = {}
        for sn in names:
            ws = wb[sn]
            raw = [list(r) for r in ws.iter_rows(values_only=True)]
            matrices[sn] = raw
            estimates.append((sn, _estimate_data_rows(raw)))

        largest_sn = max(estimates, key=lambda x: x[1])[0] if estimates else names[0]

        merged: list[dict[str, Any]] = []
        best: Optional[ParseResult] = None
        best_n = -1

        for sn in names:
            raw = matrices.get(sn) or []
            use_llm = (bool(use_ollama) or bool(anthropic_api_key.strip())) and sn == largest_sn
            sub = _parse_sheet_matrix(
                raw,
                sheet_name=sn,
                use_llm=use_llm,
                allow_ollama=bool(use_ollama),
                ollama_base_url=ollama_base_url,
                ollama_model=ollama_model,
                anthropic_api_key=anthropic_api_key,
                haiku_model=haiku_model,
                timeout=timeout,
                ollama_usage=ollama_usage,
                anthropic_usage=anthropic_usage,
            )
            merged.extend(sub.rows)
            if len(sub.rows) > best_n:
                best_n = len(sub.rows)
                best = sub

        if best is None:
            best = ParseResult(
                headers=[],
                header_row_1based=1,
                mapping={k: None for k in SEMANTIC_FIELDS},
                rows=[],
                detection_method="heuristic",
                sheets_processed=names,
            )

        return ParseResult(
            headers=best.headers,
            header_row_1based=best.header_row_1based,
            mapping=best.mapping,
            rows=merged,
            detection_method=best.detection_method,
            sheets_processed=names,
        )
    finally:
        wb.close()


def _parse_csv_bytes(
    data: bytes,
    *,
    sheet_name: str,
    use_llm: bool,
    allow_ollama: bool,
    ollama_base_url: str,
    ollama_model: str,
    anthropic_api_key: str,
    haiku_model: str,
    timeout: float,
    ollama_usage: Optional[OllamaTokenLedger] = None,
    anthropic_usage: Optional[AnthropicUsageLedger] = None,
) -> ParseResult:
    text = data.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample)
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    raw: list[list[Any]] = []
    for row in reader:
        raw.append(row)
    return _parse_sheet_matrix(
        raw,
        sheet_name=sheet_name,
        use_llm=use_llm,
        allow_ollama=allow_ollama,
        ollama_base_url=ollama_base_url,
        ollama_model=ollama_model,
        anthropic_api_key=anthropic_api_key,
        haiku_model=haiku_model,
        timeout=timeout,
        ollama_usage=ollama_usage,
        anthropic_usage=anthropic_usage,
    )


def parse_workbook_bytes(
    xlsx_bytes: bytes,
    *,
    ollama_base_url: str = "http://127.0.0.1:11434",
    ollama_model: str = "gemma3:27b",
    use_ollama: bool = True,
    timeout: float = 240.0,
    anthropic_api_key: str = "",
    haiku_model: str = "claude-haiku-4-5-20251001",
    ollama_usage: Optional[OllamaTokenLedger] = None,
    anthropic_usage: Optional[AnthropicUsageLedger] = None,
) -> ParseResult:
    return parse_uploaded_file(
        xlsx_bytes,
        "workbook.xlsx",
        ollama_base_url=ollama_base_url,
        ollama_model=ollama_model,
        use_ollama=use_ollama,
        timeout=timeout,
        anthropic_api_key=anthropic_api_key,
        haiku_model=haiku_model,
        ollama_usage=ollama_usage,
        anthropic_usage=anthropic_usage,
    )


def filter_rows_with_lookup(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        r
        for r in rows
        if r.get("_asin") or r.get("_gtin") or r.get("_sku") or r.get("_mpn")
    ]


def filter_rows_with_asin(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return filter_rows_with_lookup(rows)
