"""
Excel output: preserve source columns + appended ASIN, confidence, and optional trace column
(optional multi-sheet).
"""

from __future__ import annotations

from io import BytesIO
from typing import Any, Mapping, Sequence, Union

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def _flatten_for_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _autosize_columns(ws: Any, col_count: int, max_width: int = 60) -> None:
    for i in range(1, col_count + 1):
        letter = get_column_letter(i)
        best = 10
        for cell in ws[letter]:
            if cell.value is None:
                continue
            best = min(max_width, max(best, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = best


def passthrough_headers(
    col_order: list[str],
) -> tuple[list[str], str, str, str, str, str, str, str, str, str]:
    """
    Original column order unchanged; return
    (full_row_headers, asin_header, confidence_header, average_price_header,
    buy_box_incl_shipping_header, take_home_profit_header, roi_header,
    monthly_sales_quantity_header, trace_header, rejected_asin_header).
    ``Take home profit`` and ``ROI`` use the mapped vendor **cost** column (see parser)
    plus Keepa Buy Box landed price for HIGH/MEDIUM rows.
    """
    lower = {str(h).strip().lower() for h in col_order if h}
    asin_h = "ASIN"
    if asin_h.lower() in lower:
        asin_h = "Resolved ASIN"
    conf_h = "confidence"
    if conf_h.lower() in lower:
        conf_h = "Trasco confidence"
    avg_h = "Average price"
    if avg_h.lower() in lower:
        avg_h = "Trasco average price (6 mo)"
    if avg_h.lower() in lower:
        avg_h = "_trasco_avg_price_6m"
    bb_h = "Buy Box Price (incl. shipping if FBM)"
    if bb_h.lower() in lower:
        bb_h = "Trasco buy box price (incl. shipping if FBM)"
    if bb_h.lower() in lower:
        bb_h = "_trasco_buy_box_incl_ship"
    take_h = "Take home profit"
    if take_h.lower() in lower:
        take_h = "Trasco take home profit"
    if take_h.lower() in lower:
        take_h = "_trasco_take_home_profit"
    roi_h = "ROI"
    if roi_h.lower() in lower:
        roi_h = "Trasco ROI"
    if roi_h.lower() in lower:
        roi_h = "_trasco_roi_pct"
    msq_h = "Monthly sales quantity"
    if msq_h.lower() in lower:
        msq_h = "Trasco monthly sales quantity"
    if msq_h.lower() in lower:
        msq_h = "_trasco_monthly_sales_qty"
    log_h = "Trasco trace"
    if log_h.lower() in lower:
        log_h = "Trasco diagnostics"
    if log_h.lower() in lower:
        log_h = "_trasco_trace"
    rej_h = "Rejected ASIN (LLM)"
    if rej_h.lower() in lower:
        rej_h = "Trasco rejected ASIN"
    if rej_h.lower() in lower:
        rej_h = "_trasco_llm_rejected_asin"
    full = list(col_order) + [asin_h, conf_h, avg_h, bb_h, take_h, roi_h, msq_h, log_h, rej_h]
    return full, asin_h, conf_h, avg_h, bb_h, take_h, roi_h, msq_h, log_h, rej_h


def workbook_from_sheet_sections(
    sections: list[
        Union[
            tuple[str, list[str], list[Mapping[str, Any]]],
            tuple[str, list[str], list[Mapping[str, Any]], Any],
        ]
    ],
) -> BytesIO:
    """
    Each section: (worksheet_title, headers_in_order, row_dicts) or
    (title, headers, rows, meta) — meta is ignored (used by the pipeline only).
    Row dicts must contain exactly the keys in headers_in_order.
    """
    wb = Workbook()
    if not sections:
        ws = wb.active
        ws.title = "results"
        ws.append(["(no rows)"])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    header_font = Font(bold=True)
    fills = {
        "HIGH": PatternFill("solid", fgColor="C6EFCE"),
        "MEDIUM": PatternFill("solid", fgColor="FFEB9C"),
        "LOW": PatternFill("solid", fgColor="FFC7CE"),
        "NOT FOUND": PatternFill("solid", fgColor="D9D9D9"),
        "NOT FOUND (LLM)": PatternFill("solid", fgColor="D9D9D9"),
        "NOT FOUND (Foreign)": PatternFill("solid", fgColor="D9D9D9"),
    }
    used_titles: set[str] = set()

    def _unique_sheet_title(raw: str) -> str:
        base = (raw or "Sheet")[:31] or "Sheet"
        t = base
        n = 1
        while t in used_titles:
            n += 1
            suffix = f"_{n}"
            t = (base[: 31 - len(suffix)] + suffix)[:31]
        used_titles.add(t)
        return t

    for idx, entry in enumerate(sections):
        if len(entry) == 4:
            title, headers, rows, _meta = entry
        else:
            title, headers, rows = entry  # type: ignore[misc]
        safe_title = _unique_sheet_title(title)
        if idx == 0:
            ws = wb.active
            ws.title = safe_title
        else:
            ws = wb.create_sheet(title=safe_title)

        if not headers:
            ws.append(["(empty)"])
            continue

        ws.append(headers)
        for col_i, _h in enumerate(headers, start=1):
            ws.cell(row=1, column=col_i).font = header_font

        conf_headers = {"confidence", "Trasco confidence"}
        conf_col_idx = None
        for i, h in enumerate(headers, start=1):
            if h in conf_headers:
                conf_col_idx = i
                break

        for row in rows:
            ws.append([_flatten_for_cell(row.get(h)) for h in headers])

        if conf_col_idx:
            letter = get_column_letter(conf_col_idx)
            for r in range(2, ws.max_row + 1):
                val = ws[f"{letter}{r}"].value
                if val in fills:
                    ws[f"{letter}{r}"].fill = fills[str(val)]

        ws.freeze_panes = "A2"
        _autosize_columns(ws, len(headers))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def rows_to_workbook(rows: Sequence[Mapping[str, Any]]) -> BytesIO:
    """Backward-compatible single-sheet writer when rows already include final headers."""
    if not rows:
        return workbook_from_sheet_sections([("results", [], [], None)])
    headers = list(rows[0].keys())
    return workbook_from_sheet_sections([("results", headers, list(rows), None)])
